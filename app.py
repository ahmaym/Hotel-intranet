from flask import Flask, render_template, request, redirect, session, url_for, flash, jsonify, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from ldap3 import Server, Connection, ALL, NTLM, SIMPLE
import sqlite3
from functools import wraps
from datetime import datetime
from flask_socketio import SocketIO, emit, join_room
from os import path
from flask import session
import logging
from werkzeug.utils import secure_filename
import openpyxl
import os
from dotenv import load_dotenv
from chatbot import get_chatbot

# Load environment variables from .env file
load_dotenv()



# =================== CONFIG ====================
APP_DB = os.getenv('DATABASE_PATH', 'database.db')

app = Flask(__name__)

app.secret_key = os.getenv('SECRET_KEY', 'change_this_secret_to_something_secure')
socketio = SocketIO(app, async_mode='threading', cors_allowed_origins="*")

# File upload configuration
app.config['MAX_CONTENT_LENGTH'] = int(os.getenv('MAX_CONTENT_LENGTH', 16777216))  # 16MB default
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
UPLOAD_FOLDER = os.getenv('UPLOAD_FOLDER', 'uploads')

# Active Directory settings
app.config['LDAP_HOST'] = os.getenv('LDAP_HOST', 'your_ldap_server_ip')
app.config['LDAP_PORT'] = int(os.getenv('LDAP_PORT', 389))
app.config['LDAP_USE_SSL'] = os.getenv('LDAP_USE_SSL', 'False').lower() == 'true'
app.config['LDAP_BASE_DN'] = os.getenv('LDAP_BASE_DN', 'DC=YOUR-DOMAIN,DC=COM')
app.config['LDAP_DOMAIN'] = os.getenv('LDAP_DOMAIN', 'YOUR-DOMAIN')

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# =================== DB ====================
def get_db():
    conn = sqlite3.connect(APP_DB)
    conn.row_factory = sqlite3.Row
    return conn

def _table_exists(c, name):
    c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (name,))
    return bool(c.fetchone())

def _column_exists(c, table, column):
    c.execute(f"PRAGMA table_info({table})")
    cols = [row[1] for row in c.fetchall()]
    return column in cols

def ensure_db_schema():
    if not path.exists(APP_DB):
        return
    conn = get_db()
    c = conn.cursor()

    if _table_exists(c, 'tickets') and not _column_exists(c, 'tickets', 'ticket_number'):
        c.execute("ALTER TABLE tickets ADD COLUMN ticket_number TEXT")
        conn.commit()

    if _table_exists(c, 'tickets') and _column_exists(c, 'tickets', 'ticket_number'):
        c.execute("SELECT id FROM tickets WHERE ticket_number IS NULL OR ticket_number = ''")
        missing = c.fetchall()
        for row in missing:
            tid = row['id']
            c.execute("UPDATE tickets SET ticket_number = ? WHERE id = ?", (f"TKT-{tid:06d}", tid))
        conn.commit()
        c.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_tickets_ticket_number ON tickets(ticket_number)")
        conn.commit()

    # Add assigned_to_user and assigned_at columns for ticket assignment
    if _table_exists(c, 'tickets') and not _column_exists(c, 'tickets', 'assigned_to_user'):
        c.execute("ALTER TABLE tickets ADD COLUMN assigned_to_user INTEGER")
        conn.commit()
    
    if _table_exists(c, 'tickets') and not _column_exists(c, 'tickets', 'assigned_to_username'):
        c.execute("ALTER TABLE tickets ADD COLUMN assigned_to_username TEXT")
        conn.commit()
    
    if _table_exists(c, 'tickets') and not _column_exists(c, 'tickets', 'assigned_at'):
        c.execute("ALTER TABLE tickets ADD COLUMN assigned_at TEXT")
        conn.commit()

    c.execute("""
        CREATE TABLE IF NOT EXISTS daily_updates (
            id INTEGER PRIMARY KEY,
            image_filename TEXT,
            uploaded_at TEXT,
            uploaded_by INTEGER
        )
    """)
    conn.commit()
    c.execute("SELECT id FROM daily_updates WHERE id = 1")
    if not c.fetchone():
        c.execute("INSERT INTO daily_updates (id, image_filename, uploaded_at, uploaded_by) VALUES (1, NULL, NULL, NULL)")
        conn.commit()

    c.execute("""
        CREATE TABLE IF NOT EXISTS ticket_comments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ticket_id INTEGER NOT NULL,
            user_id TEXT,
            username TEXT NOT NULL,
            comment TEXT,
            status_change TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (ticket_id) REFERENCES tickets(id)
        )
    """)
    conn.commit()

    # Create notifications table for ticket assignments
    c.execute("""
        CREATE TABLE IF NOT EXISTS notifications (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id TEXT NOT NULL,
            username TEXT NOT NULL,
            notification_type TEXT NOT NULL,
            title TEXT NOT NULL,
            message TEXT NOT NULL,
            ticket_id INTEGER,
            is_read INTEGER DEFAULT 0,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (ticket_id) REFERENCES tickets(id)
        )
    """)
    conn.commit()

    # Create user_tasks table for My Tasks feature
    c.execute("""
        CREATE TABLE IF NOT EXISTS user_tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id TEXT NOT NULL,
            task_text TEXT NOT NULL,
            is_completed INTEGER DEFAULT 0,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            completed_at TEXT
        )
    """)
    conn.commit()

    # Add media_filename column to posts table for image attachments
    if _table_exists(c, 'posts') and not _column_exists(c, 'posts', 'media_filename'):
        c.execute("ALTER TABLE posts ADD COLUMN media_filename TEXT")
        conn.commit()

    # Create calendar_events table for dashboard calendar
    c.execute("""
        CREATE TABLE IF NOT EXISTS calendar_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id TEXT NOT NULL,
            title TEXT NOT NULL,
            description TEXT,
            event_date TEXT NOT NULL,
            event_type TEXT DEFAULT 'event',
            color TEXT DEFAULT 'sky',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()

    conn.close()

ensure_db_schema()

# Image extensions from environment variable
IMAGE_EXTENSIONS = set(os.getenv('IMAGE_EXTENSIONS', 'png,jpg,jpeg,gif,bmp,webp').split(','))
DAILY_UPDATES_FOLDER = os.getenv('DAILY_UPDATES_FOLDER', 'static/daily_updates')

def allowed_image_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in IMAGE_EXTENSIONS

def parse_features(s):
    if not s:
        return []
    return [x.strip().lower() for x in s.split(',') if x.strip()]

def normalize_department(dept):
    """Normalize department name from AD to standard format (IT, HR, ENG)"""
    if not dept:
        return ""
    dept_upper = dept.upper().strip()
    # Map common AD department names to standard codes
    if 'IT' in dept_upper or 'INFORMATION TECHNOLOGY' in dept_upper or 'TECH' in dept_upper:
        return 'IT'
    elif 'HR' in dept_upper or 'HUMAN RESOURCE' in dept_upper:
        return 'HR'
    elif 'ENG' in dept_upper or 'ENGINEER' in dept_upper or 'ENGINEERING' in dept_upper:
        return 'ENG'
    return dept.strip()

def get_time_elapsed(created_at):
    """Calculate how long ago a ticket was created in a human-readable format"""
    try:
        created = datetime.strptime(created_at, '%Y-%m-%d %H:%M:%S')
        now = datetime.now()
        delta = now - created
        
        # Handle negative deltas (future dates or timezone issues)
        total_seconds = delta.total_seconds()
        if total_seconds < 0:
            return "Just now"
        
        # Calculate time components from total seconds
        total_minutes = int(total_seconds // 60)
        total_hours = int(total_seconds // 3600)
        days = delta.days
        
        # Return human-readable format
        if days > 0:
            if days == 1:
                return "1 day"
            return f"{days} days"
        elif total_hours > 0:
            if total_hours == 1:
                return "1 hour"
            return f"{total_hours} hours"
        elif total_minutes > 0:
            if total_minutes == 1:
                return "1 minute"
            return f"{total_minutes} minutes"
        else:
            return "Just now"
    except:
        return "Unknown"

def get_time_elapsed_detailed(created_at):
    """Calculate detailed time elapsed for display in dedicated column"""
    try:
        created = datetime.strptime(created_at, '%Y-%m-%d %H:%M:%S')
        now = datetime.now()
        delta = now - created
        
        total_seconds = delta.total_seconds()
        if total_seconds < 0:
            return "0m"
        
        days = delta.days
        hours = int((total_seconds % 86400) // 3600)
        minutes = int((total_seconds % 3600) // 60)
        
        parts = []
        if days > 0:
            parts.append(f"{days}d")
        if hours > 0:
            parts.append(f"{hours}h")
        if minutes > 0 and days == 0:  # Only show minutes if less than a day
            parts.append(f"{minutes}m")
        
        if not parts:
            return "< 1m"
        
        return " ".join(parts)
    except:
        return "Unknown"

# Register the functions as Jinja2 globals
app.jinja_env.globals.update(get_time_elapsed=get_time_elapsed)
app.jinja_env.globals.update(get_time_elapsed_detailed=get_time_elapsed_detailed)

# =================== LDAP / AD AUTH ====================
def authenticate_ldap_user(username, password):
    """
    Authenticate user with Active Directory and read Level from HRCW OU or group.
    Returns (department, role, display_name, features) on success, or None on failure.
    """
    from ldap3 import Server, Connection, ALL, SIMPLE

    LDAP_HOST = app.config.get('LDAP_HOST')
    LDAP_PORT = app.config.get('LDAP_PORT', 389)
    USE_SSL = app.config.get('LDAP_USE_SSL', False)
    BASE_DN = app.config.get('LDAP_BASE_DN')
    DOMAIN = app.config.get('LDAP_DOMAIN')

    # Use UPN format for simple bind
    user_bind = f"{username}@{DOMAIN}.COM" if not username.endswith(f"@{DOMAIN}.COM") else username

    try:
        server = Server(LDAP_HOST, port=LDAP_PORT, use_ssl=USE_SSL, get_info=ALL)
        conn = Connection(server, user=user_bind, password=password, authentication=SIMPLE, receive_timeout=10)
        if not conn.bind():
            # bind failed
            print("[LDAP] bind failed:", conn.result)
            return None

        # search for the user entry to get distinguishedName, memberOf, department, givenName, sn
        conn.search(
            search_base=BASE_DN,
            search_filter=f'(sAMAccountName={username})',
            attributes=['distinguishedName', 'memberOf', 'department', 'givenName', 'sn', 'displayName']
        )

        if not conn.entries:
            conn.unbind()
            return None

        entry = conn.entries[0]
        dn = str(entry.distinguishedName) if 'distinguishedName' in entry else ''
        member_of = str(entry.memberOf) if 'memberOf' in entry else ''
        department = str(entry.department) if 'department' in entry and entry.department else 'General'
        
        # Build display name from givenName (first name) and sn (last name)
        given_name = str(entry.givenName) if 'givenName' in entry and entry.givenName else ''
        surname = str(entry.sn) if 'sn' in entry and entry.sn else ''
        if given_name and surname:
            display_name = f"{given_name} {surname}"
        elif given_name:
            display_name = given_name
        elif surname:
            display_name = surname
        else:
            # Fallback to displayName or formatted username
            display_name = str(entry.displayName) if 'displayName' in entry and entry.displayName else ' '.join(word.capitalize() for word in username.replace('.', ' ').replace('_', ' ').split())

        # Determine role & features based on Level in DN or memberOf
        role = 'viewer'
        features = 'read'

        # Prefer detecting Level0..Level4 in DN first, then memberOf
        if 'OU=Level0' in dn or 'Level0' in member_of:
            role = 'admin'
            features = 'manage_users,read,posts,announcements,edit,delete'
        elif 'OU=Level1' in dn or 'Level1' in member_of:
            role = 'manager'
            features = 'read,posts,announcements,edit'
        elif 'OU=Level2' in dn or 'Level2' in member_of:
            role = 'supervisor'
            features = 'read,posts'
        elif 'OU=Level3' in dn or 'Level3' in member_of:
            role = 'staff'
            features = 'read'
        elif 'OU=Level4' in dn or 'Level4' in member_of:
            role = 'viewer'
            features = 'read'

        conn.unbind()
        return (department, role, display_name, features)
    except Exception as e:
        print("[LDAP] Exception:", e)
        return None

# =================== DECORATORS ====================
def login_required(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return wrapped

def admin_required(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if 'user_id' not in session or session.get('role') != 'admin':
            flash('Admin privileges are required to access this page.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return wrapped
# =================== Alert ====================

@app.route('/add_alert', methods=['GET', 'POST'])
@admin_required
def add_alert():
    if request.method == 'POST':
        title = request.form['title']
        message = request.form['message']

        conn = get_db()
        c = conn.cursor()
        c.execute("INSERT INTO announcements (title, message, created_at) VALUES (?, ?, datetime('now'))", (title, message))
        conn.commit()
        conn.close()

        flash('Announcement added successfully.', 'success')
        return redirect(url_for('dashboard'))

    return render_template('add_alert.html')

# =================== AUTH ====================
@app.route('/')
def root():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        # المحاولة الأولى: التحقق من الـ Active Directory
                # المحاولة الأولى: التحقق من الـ Active DIRECTORY
        ldap_result = authenticate_ldap_user(username, password)
        if ldap_result:
            department, role, display_name, features = ldap_result
            session['user_id'] = username  # 👈 استخدم الاسم الحقيقي من الـAD
            session['username'] = username
            session['display_name'] = display_name  # Store display name for UI
            session['role'] = role
            session['department'] = department
            session['features'] = features
            session['password'] = password  # Store password for AD queries in chat

            flash('Logged in successfully via Active Directory.', 'success')
            return redirect(url_for('dashboard'))

        # لو فشل الـ AD login، نرجع نجرب قاعدة البيانات المحلية
        conn = get_db()
        c = conn.cursor()
        c.execute("SELECT id, username, password, role, features, department FROM users WHERE username = ?", (username,))
        row = c.fetchone()
        conn.close()
        if row and check_password_hash(row['password'], password):
            session['user_id'] = row['id']
            session['username'] = row['username']
            # Format username as display name (ahmed.ayman -> Ahmed Ayman)
            session['display_name'] = ' '.join(word.capitalize() for word in row['username'].replace('.', ' ').replace('_', ' ').split())
            session['role'] = row['role']
            session['features'] = row['features'] or ''
            session['department'] = row['department'] or ''
            flash('Logged in successfully (local).', 'success')
            return redirect(url_for('dashboard'))
        else:
            error = 'Wrong username or password'


    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# =================== DASHBOARD ====================
POSTS_MEDIA_FOLDER = 'static/posts_media'

@app.route('/dashboard', methods=['GET','POST'])
@login_required
def dashboard():
    conn = get_db()
    c = conn.cursor()
    feats = parse_features(session.get('features', ''))

    if request.method == 'POST':
        ptype = request.form.get('type', 'post')
        if session.get('role') != 'admin':
            if ptype == 'post' and 'posts' not in feats:
                flash('You are not allowed to post.', 'danger')
                return redirect(url_for('dashboard'))
            if ptype == 'announcement' and 'announcements' not in feats:
                flash('You are not allowed to publish announcements.', 'danger')
                return redirect(url_for('dashboard'))

        title = request.form.get('title', '').strip()
        content = request.form.get('content', '').strip()
        if not title or not content:
            flash('Title and content are required.', 'warning')
        else:
            # Handle media file upload
            media_filename = None
            media_file = request.files.get('media')
            if media_file and media_file.filename:
                if allowed_image_file(media_file.filename):
                    os.makedirs(path.join(app.root_path, POSTS_MEDIA_FOLDER), exist_ok=True)
                    ext = media_file.filename.rsplit('.', 1)[1].lower()
                    media_filename = f"post_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{session['user_id']}.{ext}"
                    save_path = path.join(app.root_path, POSTS_MEDIA_FOLDER, media_filename)
                    media_file.save(save_path)
                else:
                    flash('Unsupported image type. Allowed: png, jpg, jpeg, gif, webp.', 'warning')
                    return redirect(url_for('dashboard'))
            
            c.execute("""
                INSERT INTO posts (title, content, type, author_id, author_username, created_at, media_filename)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (title, content, ptype, session['user_id'], session['username'],
                  datetime.now().strftime('%Y-%m-%d %H:%M:%S'), media_filename))
            conn.commit()
            flash('Post published successfully.', 'success')
            return redirect(url_for('dashboard'))

    c.execute("SELECT id, title, content, type, author_id, author_username, created_at, media_filename FROM posts WHERE type = 'announcement' ORDER BY created_at DESC LIMIT 5")
    announcements = c.fetchall()

    c.execute("SELECT id, title, content, type, author_id, author_username, created_at, media_filename FROM posts WHERE type = 'post' ORDER BY created_at DESC LIMIT 20")
    posts = c.fetchall()

    today_md = datetime.now().strftime('%m-%d')
    try:
        c.execute("""
            SELECT name, job_title 
            FROM staff 
            WHERE birthday IS NOT NULL 
              AND strftime('%m-%d', birthday) = ?
        """, (today_md,))
        birthdays = c.fetchall()
    except sqlite3.OperationalError:
        birthdays = []

    c.execute("SELECT image_filename FROM daily_updates WHERE id = 1")
    daily_update_row = c.fetchone()
    daily_update_image = daily_update_row['image_filename'] if daily_update_row else None

    conn.close()
    return render_template(
        'dashboard.html',
        announcements=announcements,
        posts=posts,
        birthdays=birthdays,
        daily_update_image=daily_update_image,
        features_list=feats,
        role=session.get('role')
    )


@app.route('/daily_update/upload', methods=['POST'])
@login_required
def upload_daily_update():
    if session.get('role') != 'admin' and 'edit' not in parse_features(session.get('features', '')):
        flash('You do not have permission to update Daily Updates.', 'danger')
        return redirect(url_for('dashboard'))

    file = request.files.get('image')
    if not file or not file.filename:
        flash('Please choose an image to upload.', 'warning')
        return redirect(url_for('dashboard'))

    if not allowed_image_file(file.filename):
        flash('Unsupported image type. Allowed: png, jpg, jpeg, gif, webp.', 'warning')
        return redirect(url_for('dashboard'))

    os.makedirs(path.join(app.root_path, DAILY_UPDATES_FOLDER), exist_ok=True)
    ext = file.filename.rsplit('.', 1)[1].lower()
    new_filename = f"daily_update_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{ext}"
    save_path = path.join(app.root_path, DAILY_UPDATES_FOLDER, new_filename)

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT image_filename FROM daily_updates WHERE id = 1")
    row = c.fetchone()
    old_filename = row['image_filename'] if row else None

    if old_filename:
        old_path = path.join(app.root_path, DAILY_UPDATES_FOLDER, old_filename)
        try:
            if path.exists(old_path):
                os.remove(old_path)
        except Exception:
            pass

    file.save(save_path)
    c.execute(
        "UPDATE daily_updates SET image_filename = ?, uploaded_at = ?, uploaded_by = ? WHERE id = 1",
        (new_filename, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), session.get('user_id'))
    )
    conn.commit()
    conn.close()

    flash('Daily Updates image uploaded successfully.', 'success')
    return redirect(url_for('dashboard'))


@app.route('/daily_update/delete', methods=['POST'])
@login_required
def delete_daily_update():
    if session.get('role') != 'admin' and 'edit' not in parse_features(session.get('features', '')):
        flash('You do not have permission to update Daily Updates.', 'danger')
        return redirect(url_for('dashboard'))

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT image_filename FROM daily_updates WHERE id = 1")
    row = c.fetchone()
    old_filename = row['image_filename'] if row else None

    if old_filename:
        old_path = path.join(app.root_path, DAILY_UPDATES_FOLDER, old_filename)
        try:
            if path.exists(old_path):
                os.remove(old_path)
        except Exception:
            pass

    c.execute("UPDATE daily_updates SET image_filename = NULL, uploaded_at = NULL, uploaded_by = NULL WHERE id = 1")
    conn.commit()
    conn.close()

    flash('Daily Updates image removed.', 'success')
    return redirect(url_for('dashboard'))
# =================== Posts and announcement ====================

@app.route('/delete_announcement/<int:ann_id>', methods=['POST'])
@login_required
def delete_announcement(ann_id):
    conn = get_db()
    c = conn.cursor()
    
    # Check ownership
    c.execute("SELECT author_id FROM posts WHERE id = ? AND type = 'announcement'", (ann_id,))
    post = c.fetchone()
    
    if not post:
        conn.close()
        flash('Announcement not found.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Only the author or admin can delete
    if post['author_id'] != session.get('user_id') and session.get('role') != 'admin':
        conn.close()
        flash('You do not have permission to delete this announcement.', 'danger')
        return redirect(url_for('dashboard'))
    
    c.execute("DELETE FROM posts WHERE id = ?", (ann_id,))
    conn.commit()
    conn.close()

    flash('Announcement deleted successfully.', 'success')
    return redirect(url_for('dashboard'))




@app.route('/delete_post/<int:post_id>', methods=['POST'])
@login_required
def delete_post(post_id):
    conn = get_db()
    c = conn.cursor()
    
    # Check ownership
    c.execute("SELECT author_id FROM posts WHERE id = ? AND type = 'post'", (post_id,))
    post = c.fetchone()
    
    if not post:
        conn.close()
        flash('Post not found.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Only the author or admin can delete
    if post['author_id'] != session.get('user_id') and session.get('role') != 'admin':
        conn.close()
        flash('You do not have permission to delete this post.', 'danger')
        return redirect(url_for('dashboard'))
    
    c.execute("DELETE FROM posts WHERE id = ?", (post_id,))
    conn.commit()
    conn.close()

    flash('Post deleted successfully.', 'success')
    return redirect(url_for('dashboard'))


@app.route('/edit_announcement/<int:ann_id>', methods=['GET', 'POST'])
@login_required
def edit_announcement(ann_id):
    conn = get_db()
    c = conn.cursor()
    
    # Get the announcement
    c.execute("SELECT * FROM posts WHERE id = ? AND type = 'announcement'", (ann_id,))
    announcement = c.fetchone()
    
    if not announcement:
        conn.close()
        flash('Announcement not found.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Only the author can edit (not even admin)
    if announcement['author_id'] != session.get('user_id'):
        conn.close()
        flash('You do not have permission to edit this announcement.', 'danger')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        content = request.form.get('content', '').strip()
        
        if not title or not content:
            flash('Title and content are required.', 'warning')
        else:
            c.execute("""
                UPDATE posts 
                SET title = ?, content = ? 
                WHERE id = ?
            """, (title, content, ann_id))
            conn.commit()
            conn.close()
            flash('Announcement updated successfully.', 'success')
            return redirect(url_for('dashboard'))
    
    conn.close()
    return render_template('edit_post.html', post=announcement, post_type='announcement')


@app.route('/edit_post/<int:post_id>', methods=['GET', 'POST'])
@login_required
def edit_post(post_id):
    conn = get_db()
    c = conn.cursor()
    
    # Get the post
    c.execute("SELECT * FROM posts WHERE id = ? AND type = 'post'", (post_id,))
    post = c.fetchone()
    
    if not post:
        conn.close()
        flash('Post not found.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Only the author can edit (not even admin)
    if post['author_id'] != session.get('user_id'):
        conn.close()
        flash('You do not have permission to edit this post.', 'danger')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        content = request.form.get('content', '').strip()
        
        if not title or not content:
            flash('Title and content are required.', 'warning')
        else:
            c.execute("""
                UPDATE posts 
                SET title = ?, content = ? 
                WHERE id = ?
            """, (title, content, post_id))
            conn.commit()
            conn.close()
            flash('Post updated successfully.', 'success')
            return redirect(url_for('dashboard'))
    
    conn.close()
    return render_template('edit_post.html', post=post, post_type='post')


# =================== USERS ====================
@app.route('/users')
@admin_required
def users_list():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id, username, role, department, features, created_at FROM users ORDER BY id")
    users = c.fetchall()
    conn.close()
    return render_template('users.html', users=users)

@app.route('/add_user', methods=['GET', 'POST'])
@admin_required
def add_user():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        role = request.form.get('role', 'user')
        features = request.form.getlist('features')
        features_str = ",".join(features)
        department = request.form.get('department', '').strip()

        if not username or not password:
            flash('Please fill in the required fields.', 'warning')
            return redirect(url_for('add_user'))

        hashed = generate_password_hash(password)
        conn = get_db()
        c = conn.cursor()
        try:
            c.execute("""
                INSERT INTO users (username, password, role, department, features)
                VALUES (?, ?, ?, ?, ?)
            """, (username, hashed, role, department, features_str))
            conn.commit()
            flash('User added successfully.', 'success')
            return redirect(url_for('users_list'))
        except sqlite3.IntegrityError:
            flash('Username already exists.', 'danger')
        finally:
            conn.close()

    feature_options = ['read', 'posts', 'announcements']  # للاختيار
    return render_template('add_user.html', feature_options=feature_options)

@app.route('/edit_user/<int:user_id>', methods=['GET', 'POST'])
@admin_required
def edit_user(user_id):
    conn = get_db()
    c = conn.cursor()
    
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        role = request.form.get('role', 'user')
        department = request.form.get('department', '').strip()
        features = request.form.getlist('features')
        features_str = ",".join(features)

        if password:
            hashed = generate_password_hash(password)
            c.execute("""UPDATE users 
                         SET username=?, password=?, role=?, department=?, features=? 
                         WHERE id=?""",
                      (username, hashed, role, department, features_str, user_id))
        else:
            c.execute("""UPDATE users 
                         SET username=?, role=?, department=?, features=? 
                         WHERE id=?""",
                      (username, role, department, features_str, user_id))
        conn.commit()
        conn.close()
        flash("User updated successfully.", "success")
        return redirect(url_for('users_list'))
    else:
        c.execute("SELECT * FROM users WHERE id=?", (user_id,))
        user = c.fetchone()
        conn.close()
        if not user:
            flash("User not found.", "danger")
            return redirect(url_for('users_list'))
        return render_template('edit_user.html', user=user)


@app.route('/delete_user/<int:user_id>', methods=['POST'])
@admin_required
def delete_user(user_id):
    if user_id == session.get('user_id'):
        flash('You cannot delete yourself.', 'warning'); return redirect(url_for('users_list'))
    conn = get_db(); c = conn.cursor()
    c.execute("DELETE FROM users WHERE id = ?", (user_id,)); conn.commit(); conn.close()
    flash('User deleted successfully.', 'success'); return redirect(url_for('users_list'))
# =================== STAFF ====================
@app.route('/staff')
@login_required
def staff_list():
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT id, name, job_title, department, birthday FROM staff ORDER BY name")
    staff = c.fetchall(); conn.close()
    return render_template('staff.html', staff=staff)

@app.route('/add_staff', methods=['GET','POST'])
@admin_required
def add_staff():
    if request.method == 'POST':
        name = request.form.get('name','').strip()
        job_title = request.form.get('job_title','').strip()
        birthday = request.form.get('birthday','').strip()
        department = request.form.get('department','').strip()
        conn = get_db(); c = conn.cursor()
        c.execute("INSERT INTO staff (name, job_title, department, birthday) VALUES (?, ?, ?, ?)",
                  (name, job_title, department, birthday))
        conn.commit(); conn.close(); flash(f"Employee {name} added successfully.", "success")
        return redirect(url_for('dashboard'))
    return render_template('add_staff.html')

@app.route('/import_staff', methods=['GET', 'POST'])
@admin_required
def import_staff():
    """
    Import staff members from Excel file.
    Expected columns: Name, Job Title, Department, Birthday (optional)
    """
    if request.method == 'POST':
        # Check if file is in request
        if 'file' not in request.files:
            flash('❌ No file selected. Please upload an Excel file.', 'danger')
            return redirect(url_for('import_staff'))
        
        file = request.files['file']
        
        if file.filename == '':
            flash('❌ No file selected. Please upload an Excel file.', 'danger')
            return redirect(url_for('import_staff'))
        
        if not allowed_file(file.filename):
            flash('❌ Invalid file format. Please upload an Excel file (.xlsx or .xls).', 'danger')
            return redirect(url_for('import_staff'))
        
        try:
            # Load Excel file
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            # Get headers from first row
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(cell.value.strip().lower())
            
            if not headers:
                flash('❌ Excel file is empty. Please add headers and data.', 'danger')
                return redirect(url_for('import_staff'))
            
            # Validate required columns
            required_cols = {'name', 'job title', 'department'}
            found_cols = {col for col in headers}
            
            if not required_cols.issubset(found_cols):
                missing = required_cols - found_cols
                flash(f'❌ Missing required columns: {", ".join(missing)}', 'danger')
                return redirect(url_for('import_staff'))
            
            # Find column indices
            col_map = {header: idx + 1 for idx, header in enumerate(headers)}
            
            # Process data rows
            conn = get_db()
            c = conn.cursor()
            
            added_count = 0
            error_rows = []
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Get values from columns
                    name = str(row[col_map['name'] - 1]).strip() if row[col_map['name'] - 1] else ''
                    job_title = str(row[col_map['job title'] - 1]).strip() if row[col_map['job title'] - 1] else ''
                    department = str(row[col_map['department'] - 1]).strip() if row[col_map['department'] - 1] else ''
                    birthday = str(row[col_map.get('birthday', 0) - 1]).strip() if col_map.get('birthday') and row[col_map['birthday'] - 1] else ''
                    
                    # Validate required fields
                    if not name or not job_title or not department:
                        error_rows.append(f"Row {row_idx}: Missing required fields (Name, Job Title, Department)")
                        continue
                    
                    # Check if employee already exists
                    c.execute("SELECT id FROM staff WHERE name=? AND job_title=?", (name, job_title))
                    if c.fetchone():
                        error_rows.append(f"Row {row_idx}: Employee '{name}' already exists")
                        continue
                    
                    # Insert new staff member
                    c.execute(
                        "INSERT INTO staff (name, job_title, department, birthday) VALUES (?, ?, ?, ?)",
                        (name, job_title, department, birthday if birthday else None)
                    )
                    added_count += 1
                    
                except Exception as e:
                    error_rows.append(f"Row {row_idx}: {str(e)}")
            
            # Commit changes
            conn.commit()
            conn.close()
            
            # Flash results
            if added_count > 0:
                flash(f'✅ Successfully imported {added_count} staff member(s).', 'success')
            
            if error_rows:
                error_msg = '<br>'.join(error_rows[:5])  # Show first 5 errors
                if len(error_rows) > 5:
                    error_msg += f'<br>... and {len(error_rows) - 5} more errors'
                flash(f'⚠️ Import completed with {len(error_rows)} error(s):<br>{error_msg}', 'warning')
            
            if added_count == 0 and error_rows:
                flash('❌ No staff members were imported. Please check the errors above.', 'danger')
            
            return redirect(url_for('staff_list'))
        
        except Exception as e:
            flash(f'❌ Error processing file: {str(e)}', 'danger')
            return redirect(url_for('import_staff'))
    
    return render_template('import_staff.html')

@app.route('/download_staff_template')
@admin_required
def download_staff_template():
    """Generate and download a sample Excel template for staff import"""
    from io import BytesIO
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Staff"
    
    # Add headers
    headers = ["Name", "Job Title", "Department", "Birthday"]
    ws.append(headers)
    
    # Style header row
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Add sample data
    sample_data = [
        ["احمد محمد", "مدير الفندق", "الإدارة", "1990-05-15"],
        ["فاطمة علي", "موظفة استقبال", "الاستقبال", "1995-03-20"],
        ["محمود حسن", "عامل النظافة", "الصيانة", "1988-07-10"],
    ]
    
    for row_data in sample_data:
        ws.append(row_data)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    
    # Create in-memory file
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='staff_template.xlsx'
    )

@app.route('/edit_staff/<int:staff_id>', methods=['GET', 'POST'])
@admin_required
def edit_staff(staff_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id, name, job_title, department, birthday FROM staff WHERE id=?", (staff_id,))
    staff = c.fetchone()
    if not staff:
        conn.close()
        flash("Employee not found.", "warning")
        return redirect(url_for("staff_list"))

    if request.method == 'POST':
        name = request.form.get('name','').strip()
        job_title = request.form.get('job_title','').strip()
        birthday = request.form.get('birthday','').strip()
        department = request.form.get('department','').strip()

        c.execute("""
            UPDATE staff SET name=?, job_title=?, department=?, birthday=? WHERE id=?
        """, (name, job_title, department, birthday, staff_id))
        conn.commit()
        conn.close()
        flash("Employee details updated successfully.", "success")
        return redirect(url_for("staff_list"))

    conn.close()
    return render_template("edit_staff.html", staff=staff)

@app.route('/delete_staff/<int:staff_id>', methods=['POST'])
@admin_required
def delete_staff(staff_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM staff WHERE id=?", (staff_id,))
    conn.commit()
    conn.close()
    flash("Employee deleted successfully.", "success")
    return redirect(url_for("staff_list"))
# =================== TICKETS ====================
@app.route('/tickets', methods=['GET','POST'])
@login_required
def tickets():
    conn = get_db()
    c = conn.cursor()

    if request.method == 'POST':
        title = request.form.get('title','').strip()
        description = request.form.get('description','').strip()
        dept = request.form.get('department','IT')

        if not title or not description:
            flash("Title and description are required for the ticket.", "warning")
        else:
            # Use explicit local timestamp to avoid timezone issues
            created_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            c.execute("""
                INSERT INTO tickets (title, description, created_by, assigned_department, created_at)
                VALUES (?, ?, ?, ?, ?)
            """, (title, description, session['user_id'], dept, created_at))
            conn.commit()
            ticket_id = c.lastrowid
            try:
                c.execute(
                    "UPDATE tickets SET ticket_number = ? WHERE id = ?",
                    (f"TKT-{ticket_id:06d}", ticket_id)
                )
                conn.commit()
            except Exception:
                pass
            flash(f"Ticket submitted successfully. Reference: TKT-{ticket_id:06d}", "success")

    raw_dept = session.get("department", "")
    user_dept = normalize_department(raw_dept)
    
    print(f"[TICKETS] User: {session.get('username')}, Raw Dept: {raw_dept}, Normalized: {user_dept}")
    
    q = (request.args.get('q') or '').strip()
    status_filter = (request.args.get('status') or '').strip()
    sort = (request.args.get('sort') or 'newest').strip()

    where = []
    params = []

    if user_dept in ["IT", "HR", "ENG"]:
        where.append("(t.assigned_department = ? OR t.created_by = ?)")
        params.extend([user_dept, session['user_id']])
    else:
        where.append("t.created_by = ?")
        params.append(session['user_id'])

    if status_filter in ['open', 'in_progress', 'closed']:
        where.append("t.status = ?")
        params.append(status_filter)

    if q:
        like = f"%{q}%"
        where.append("(t.ticket_number LIKE ? OR t.title LIKE ? OR t.description LIKE ? OR u.username LIKE ? OR CAST(t.created_by AS TEXT) LIKE ?)")
        params.extend([like, like, like, like, like])

    order_by = "t.created_at DESC" if sort != 'oldest' else "t.created_at ASC"
    where_sql = " AND ".join(where) if where else "1=1"

    c.execute(
        f"""SELECT t.*, u.username FROM tickets t
            LEFT JOIN users u ON t.created_by=u.id
            WHERE {where_sql}
            ORDER BY {order_by}""",
        tuple(params)
    )

    tickets = c.fetchall()
    
    # Fetch comments for each ticket
    ticket_comments = {}
    for ticket in tickets:
        c.execute("""
            SELECT * FROM ticket_comments 
            WHERE ticket_id = ? 
            ORDER BY created_at ASC
        """, (ticket['id'],))
        ticket_comments[ticket['id']] = c.fetchall()
    
    conn.close()
    return render_template(
        "tickets.html",
        tickets=tickets,
        ticket_comments=ticket_comments,
        user_dept=user_dept,
        q=q,
        status_filter=status_filter,
        sort=sort,
    )

@app.route('/ticket/<int:ticket_id>/status', methods=['POST'])
@login_required
def update_ticket_status(ticket_id):
    raw_dept = session.get("department", "")
    user_dept = normalize_department(raw_dept)
    current_user = session.get('user_id', '')  # AD username
    
    # Check if user can update this ticket
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT assigned_department, assigned_to_username FROM tickets WHERE id = ?", (ticket_id,))
    ticket = c.fetchone()
    
    if not ticket:
        conn.close()
        flash("Ticket not found.", "danger")
        return redirect(url_for("tickets"))
    
    ticket_dept = ticket['assigned_department']
    assigned_to = ticket['assigned_to_username']
    
    # Permission check: Only the assigned user or Ahmed.Ayman can update status
    # If ticket is assigned, only the assigned user or Ahmed.Ayman can update
    is_super_user = current_user.lower() == 'ahmed.ayman'
    is_assigned_user = assigned_to and current_user.lower() == assigned_to.lower()
    
    if assigned_to:
        # Ticket is assigned - only assigned user or super user can update
        if not is_assigned_user and not is_super_user:
            conn.close()
            flash("Only the assigned user can update this ticket's status.", "danger")
            return redirect(url_for("tickets"))
    else:
        # Ticket is not assigned - department check applies (or super user)
        if user_dept != ticket_dept and not is_super_user:
            conn.close()
            flash("You do not have permission to update this ticket status.", "danger")
            return redirect(url_for("tickets"))

    new_status = request.form.get("status","open")
    comment = request.form.get("comment", "").strip()
    
    # Require comment when closing a ticket
    if new_status == "closed" and not comment:
        conn.close()
        flash("A comment is required when closing a ticket.", "warning")
        return redirect(url_for("tickets"))
    
    # Get current status to check if it changed
    c.execute("SELECT status FROM tickets WHERE id = ?", (ticket_id,))
    current = c.fetchone()
    old_status = current['status'] if current else 'open'
    
    # Update ticket status
    c.execute("UPDATE tickets SET status=? WHERE id=?", (new_status, ticket_id))
    conn.commit()
    
    # Log the status change
    username = session.get('username', 'Unknown')
    user_id = session.get('user_id', '')
    
    # Create status change message
    status_msg = f"updated the ticket status from {old_status.replace('_', ' ').title()} to {new_status.replace('_', ' ').title()}"
    
    c.execute("""
        INSERT INTO ticket_comments (ticket_id, user_id, username, comment, status_change, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (ticket_id, user_id, username, comment if comment else None, status_msg, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    conn.commit()
    conn.close()
    
    flash("Ticket status updated successfully.", "success")
    return redirect(url_for("tickets"))

@app.route('/ticket/<int:ticket_id>/delete', methods=['POST'])
@login_required
def delete_ticket(ticket_id):
    """Delete a single ticket - Ahmed.Ayman only"""
    current_user = session.get('user_id', '')
    
    # Only Ahmed.Ayman can delete tickets
    if current_user.lower() != 'ahmed.ayman':
        flash("You do not have permission to delete tickets.", "danger")
        return redirect(url_for("tickets"))
    
    conn = get_db()
    c = conn.cursor()
    
    # Delete ticket comments first (foreign key constraint)
    c.execute("DELETE FROM ticket_comments WHERE ticket_id = ?", (ticket_id,))
    # Delete notifications related to this ticket
    c.execute("DELETE FROM notifications WHERE ticket_id = ?", (ticket_id,))
    # Delete the ticket
    c.execute("DELETE FROM tickets WHERE id = ?", (ticket_id,))
    conn.commit()
    conn.close()
    
    flash("Ticket deleted successfully.", "success")
    return redirect(url_for("tickets"))

@app.route('/tickets/bulk-delete', methods=['POST'])
@login_required
def bulk_delete_tickets():
    """Bulk delete tickets - Ahmed.Ayman only"""
    current_user = session.get('user_id', '')
    
    # Only Ahmed.Ayman can delete tickets
    if current_user.lower() != 'ahmed.ayman':
        flash("You do not have permission to delete tickets.", "danger")
        return redirect(url_for("tickets"))
    
    ticket_ids = request.form.getlist('ticket_ids')
    
    if not ticket_ids:
        flash("No tickets selected for deletion.", "warning")
        return redirect(url_for("tickets"))
    
    conn = get_db()
    c = conn.cursor()
    
    deleted_count = 0
    for ticket_id in ticket_ids:
        try:
            tid = int(ticket_id)
            # Delete ticket comments first
            c.execute("DELETE FROM ticket_comments WHERE ticket_id = ?", (tid,))
            # Delete notifications
            c.execute("DELETE FROM notifications WHERE ticket_id = ?", (tid,))
            # Delete the ticket
            c.execute("DELETE FROM tickets WHERE id = ?", (tid,))
            deleted_count += 1
        except (ValueError, Exception) as e:
            logging.error(f"Error deleting ticket {ticket_id}: {e}")
            continue
    
    conn.commit()
    conn.close()
    
    flash(f"Successfully deleted {deleted_count} ticket(s).", "success")
    return redirect(url_for("tickets"))

@app.route('/api/ticket-users/search', methods=['GET'])
@login_required
def api_ticket_users_search():
    """API endpoint to search for users in Active Directory for ticket assignment"""
    from ldap3 import Server, Connection, ALL, SIMPLE
    
    # Check if user is logged in
    if 'user_id' not in session:
        return jsonify([]), 401
    
    search_term = request.args.get('search', '').strip()
    
    try:
        # Get stored password from session
        password = session.get('password')
        username = session.get('user_id')
        
        if not password or not username:
            return jsonify([]), 401
        
        # Connect to AD using UPN format (same as login) with SIMPLE authentication
        server = Server(app.config['LDAP_HOST'], port=app.config['LDAP_PORT'], use_ssl=app.config['LDAP_USE_SSL'], get_info=ALL)
        
        # Use UPN format like login does
        user_bind = f"{username}@{app.config['LDAP_DOMAIN']}.COM"
        
        conn = Connection(server, user=user_bind, password=password, authentication=SIMPLE)
        
        if not conn.bind():
            print(f"[TICKET-API] Failed to bind with UPN: {user_bind}")
            return jsonify([]), 401
        
        # Search filter - require search term for performance
        if search_term and len(search_term) >= 2:
            search_filter = f'(&(objectClass=user)(|(displayName=*{search_term}*)(sAMAccountName=*{search_term}*)))'
        else:
            # Return empty if no search term or too short
            conn.unbind()
            return jsonify([]), 200
        
        conn.search(
            search_base=app.config['LDAP_BASE_DN'],
            search_filter=search_filter,
            attributes=['sAMAccountName', 'displayName', 'department', 'mail']
        )
        
        users = []
        for entry in conn.entries:
            try:
                username_val = str(entry.sAMAccountName).strip() if entry.sAMAccountName else ''
                display_name_val = str(entry.displayName).strip() if entry.displayName else username_val
                department_val = str(entry.department).strip() if entry.department else 'No Department'
                
                if username_val:
                    users.append({
                        'username': username_val,
                        'display_name': display_name_val,
                        'department': department_val
                    })
            except Exception as e:
                print(f"[TICKET-API] Error processing entry: {str(e)}")
        
        conn.unbind()
        
        # Sort by display name and limit results
        users.sort(key=lambda x: x['display_name'].lower())
        return jsonify(users[:30]), 200
        
    except Exception as e:
        print(f"[TICKET-API] EXCEPTION: {type(e).__name__}: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify([]), 500

@app.route('/ticket/<int:ticket_id>/assign', methods=['POST'])
@login_required
def assign_ticket(ticket_id):
    """Assign a ticket to a specific AD user"""
    raw_dept = session.get("department", "")
    user_dept = normalize_department(raw_dept)
    current_user_id = session.get('user_id')
    
    conn = get_db()
    c = conn.cursor()
    
    # Get ticket details
    c.execute("SELECT * FROM tickets WHERE id = ?", (ticket_id,))
    ticket = c.fetchone()
    
    if not ticket:
        conn.close()
        flash("Ticket not found.", "danger")
        return redirect(url_for("tickets"))
    
    ticket_dept = ticket['assigned_department']
    ticket_creator = ticket['created_by']
    
    # Check permission: department user OR ticket creator can assign
    can_assign = (user_dept == ticket_dept) or (current_user_id == ticket_creator)
    
    if not can_assign:
        conn.close()
        flash("You do not have permission to assign this ticket.", "danger")
        return redirect(url_for("tickets"))
    
    # Get AD username directly from form (no longer using database user ID)
    assigned_username = request.form.get("assign_to_user", "").strip()
    assigned_display_name = request.form.get("assign_to_display_name", "").strip() or assigned_username
    
    if not assigned_username:
        conn.close()
        flash("Please select a user to assign the ticket to.", "warning")
        return redirect(url_for("tickets"))
    
    assigned_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Update ticket with assignment (store AD username, not database ID)
    c.execute("""
        UPDATE tickets 
        SET assigned_to_user = NULL, assigned_to_username = ?, assigned_at = ?
        WHERE id = ?
    """, (assigned_username, assigned_at, ticket_id))
    conn.commit()
    
    # Log the assignment in ticket_comments
    assigner_username = session.get('username', 'Unknown')
    assigner_id = session.get('user_id', '')
    
    assignment_msg = f"assigned the ticket to {assigned_username}"
    
    c.execute("""
        INSERT INTO ticket_comments (ticket_id, user_id, username, comment, status_change, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (ticket_id, assigner_id, assigner_username, None, assignment_msg, assigned_at))
    conn.commit()
    
    # Create notification for the assigned user
    c.execute("SELECT ticket_number, title FROM tickets WHERE id = ?", (ticket_id,))
    ticket_info = c.fetchone()
    ticket_title = ticket_info['title'] if ticket_info else 'Unknown'
    ticket_number = ticket_info['ticket_number'] if ticket_info else ''
    
    notification_title = f"Ticket Assigned: {ticket_number}"
    notification_message = f"You have been assigned to ticket '{ticket_title}' by {assigner_username}."
    
    c.execute("""
        INSERT INTO notifications (user_id, username, notification_type, title, message, ticket_id, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (assigned_username, assigned_username, 'ticket_assignment', notification_title, notification_message, ticket_id, assigned_at))
    conn.commit()
    conn.close()
    
    flash(f"Ticket assigned to {assigned_username} successfully.", "success")
    return redirect(url_for("tickets"))

@app.route('/ticket/<int:ticket_id>/unassign', methods=['POST'])
@login_required
def unassign_ticket(ticket_id):
    """Remove assignment from a ticket"""
    raw_dept = session.get("department", "")
    user_dept = normalize_department(raw_dept)
    current_user_id = session.get('user_id')
    
    conn = get_db()
    c = conn.cursor()
    
    # Get ticket details
    c.execute("SELECT * FROM tickets WHERE id = ?", (ticket_id,))
    ticket = c.fetchone()
    
    if not ticket:
        conn.close()
        flash("Ticket not found.", "danger")
        return redirect(url_for("tickets"))
    
    ticket_dept = ticket['assigned_department']
    ticket_creator = ticket['created_by']
    old_assignee = ticket['assigned_to_username']
    
    # Check permission: department user OR ticket creator can unassign
    can_unassign = (user_dept == ticket_dept) or (current_user_id == ticket_creator)
    
    if not can_unassign:
        conn.close()
        flash("You do not have permission to unassign this ticket.", "danger")
        return redirect(url_for("tickets"))
    
    # Clear assignment
    c.execute("""
        UPDATE tickets 
        SET assigned_to_user = NULL, assigned_to_username = NULL, assigned_at = NULL
        WHERE id = ?
    """, (ticket_id,))
    conn.commit()
    
    # Log the unassignment
    username = session.get('username', 'Unknown')
    user_id = session.get('user_id', '')
    unassign_msg = f"removed assignment from {old_assignee}" if old_assignee else "removed ticket assignment"
    
    c.execute("""
        INSERT INTO ticket_comments (ticket_id, user_id, username, comment, status_change, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (ticket_id, user_id, username, None, unassign_msg, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    conn.commit()
    conn.close()
    
    flash("Ticket assignment removed.", "success")
    return redirect(url_for("tickets"))

# ==================== NOTIFICATIONS API ====================

@app.route('/api/notifications', methods=['GET'])
@login_required
def get_notifications():
    """Get unread notifications for the current user"""
    username = session.get('user_id', '')  # AD username
    
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        SELECT n.*, t.ticket_number 
        FROM notifications n
        LEFT JOIN tickets t ON n.ticket_id = t.id
        WHERE n.user_id = ? AND n.is_read = 0
        ORDER BY n.created_at DESC
        LIMIT 20
    """, (username,))
    notifications = [dict(row) for row in c.fetchall()]
    conn.close()
    
    return jsonify(notifications)

@app.route('/api/notifications/<int:notification_id>/read', methods=['POST'])
@login_required
def mark_notification_read(notification_id):
    """Mark a notification as read"""
    username = session.get('user_id', '')
    
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE notifications SET is_read = 1 WHERE id = ? AND user_id = ?", (notification_id, username))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True})

@app.route('/api/notifications/read-all', methods=['POST'])
@login_required
def mark_all_notifications_read():
    """Mark all notifications as read for the current user"""
    username = session.get('user_id', '')
    
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE notifications SET is_read = 1 WHERE user_id = ?", (username,))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True})

# ==================== MY TASKS API ====================

@app.route('/api/tasks', methods=['GET'])
@login_required
def get_tasks():
    """Get all tasks for the current user"""
    user_id = session.get('user_id', '')
    
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        SELECT id, task_text, is_completed, created_at, completed_at 
        FROM user_tasks 
        WHERE user_id = ? 
        ORDER BY is_completed ASC, created_at DESC
    """, (user_id,))
    tasks = [dict(row) for row in c.fetchall()]
    conn.close()
    
    return jsonify(tasks)

@app.route('/api/tasks', methods=['POST'])
@login_required
def add_task():
    """Add a new task for the current user"""
    user_id = session.get('user_id', '')
    data = request.get_json()
    task_text = data.get('task_text', '').strip()
    
    if not task_text:
        return jsonify({'error': 'Task text is required'}), 400
    
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        INSERT INTO user_tasks (user_id, task_text, created_at)
        VALUES (?, ?, ?)
    """, (user_id, task_text, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    conn.commit()
    task_id = c.lastrowid
    conn.close()
    
    return jsonify({'id': task_id, 'task_text': task_text, 'is_completed': 0})

@app.route('/api/tasks/<int:task_id>/toggle', methods=['POST'])
@login_required
def toggle_task(task_id):
    """Toggle task completion status"""
    user_id = session.get('user_id', '')
    
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT is_completed FROM user_tasks WHERE id = ? AND user_id = ?", (task_id, user_id))
    task = c.fetchone()
    
    if not task:
        conn.close()
        return jsonify({'error': 'Task not found'}), 404
    
    new_status = 0 if task['is_completed'] else 1
    completed_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S') if new_status else None
    
    c.execute("UPDATE user_tasks SET is_completed = ?, completed_at = ? WHERE id = ? AND user_id = ?", 
              (new_status, completed_at, task_id, user_id))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'is_completed': new_status})

@app.route('/api/tasks/<int:task_id>', methods=['DELETE'])
@login_required
def delete_task(task_id):
    """Delete a task"""
    user_id = session.get('user_id', '')
    
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM user_tasks WHERE id = ? AND user_id = ?", (task_id, user_id))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True})

# ==================== CALENDAR EVENTS API ====================

@app.route('/api/calendar/events', methods=['GET'])
@login_required
def get_calendar_events():
    """Get calendar events for a specific month"""
    user_id = session.get('user_id', '')
    year = request.args.get('year', datetime.now().year, type=int)
    month = request.args.get('month', datetime.now().month, type=int)
    
    # Get first and last day of month
    first_day = f"{year}-{month:02d}-01"
    if month == 12:
        last_day = f"{year + 1}-01-01"
    else:
        last_day = f"{year}-{month + 1:02d}-01"
    
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        SELECT id, title, description, event_date, event_type, color, created_at
        FROM calendar_events 
        WHERE user_id = ? AND event_date >= ? AND event_date < ?
        ORDER BY event_date ASC
    """, (user_id, first_day, last_day))
    events = [dict(row) for row in c.fetchall()]
    conn.close()
    
    return jsonify(events)

@app.route('/api/calendar/events', methods=['POST'])
@login_required
def create_calendar_event():
    """Create a new calendar event"""
    user_id = session.get('user_id', '')
    data = request.get_json()
    
    title = data.get('title', '').strip()
    description = data.get('description', '').strip()
    event_date = data.get('event_date', '')
    event_type = data.get('event_type', 'event')
    color = data.get('color', 'sky')
    
    if not title or not event_date:
        return jsonify({'error': 'Title and date are required'}), 400
    
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        INSERT INTO calendar_events (user_id, title, description, event_date, event_type, color, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (user_id, title, description, event_date, event_type, color, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    conn.commit()
    event_id = c.lastrowid
    conn.close()
    
    return jsonify({
        'id': event_id,
        'title': title,
        'description': description,
        'event_date': event_date,
        'event_type': event_type,
        'color': color
    })

@app.route('/api/calendar/events/<int:event_id>', methods=['PUT'])
@login_required
def update_calendar_event(event_id):
    """Update a calendar event"""
    user_id = session.get('user_id', '')
    data = request.get_json()
    
    title = data.get('title', '').strip()
    description = data.get('description', '').strip()
    event_date = data.get('event_date', '')
    event_type = data.get('event_type', 'event')
    color = data.get('color', 'sky')
    
    if not title or not event_date:
        return jsonify({'error': 'Title and date are required'}), 400
    
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        UPDATE calendar_events 
        SET title = ?, description = ?, event_date = ?, event_type = ?, color = ?
        WHERE id = ? AND user_id = ?
    """, (title, description, event_date, event_type, color, event_id, user_id))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True})

@app.route('/api/calendar/events/<int:event_id>', methods=['DELETE'])
@login_required
def delete_calendar_event(event_id):
    """Delete a calendar event"""
    user_id = session.get('user_id', '')
    
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM calendar_events WHERE id = ? AND user_id = ?", (event_id, user_id))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True})

# ==================== CHAT SYSTEM (ACTIVE DIRECTORY) ====================

from ldap3 import Server, Connection, ALL, NTLM
from flask_socketio import SocketIO, emit, join_room

socketio = SocketIO(app, async_mode='threading', cors_allowed_origins="*")
connected_users = {}

# Add debug logs to track user fetching and message delivery
logging.basicConfig(level=logging.DEBUG)

# Function to fetch users from Active Directory
def get_ad_users(search_query=None):
    """Fetch users from Active Directory using UPN format (same as login)"""
    users = []
    try:
        logging.debug(f"[AD] Starting user fetch with search_query: '{search_query}'")
        
        # Check session
        if 'user_id' not in session:
            logging.error("[AD] User ID not in session")
            return []
        if 'password' not in session:
            logging.error("[AD] Password not in session")
            return []
        
        user_id = session.get('user_id')
        password = session.get('password')
        
        logging.debug(f"[AD] Using credentials - User: {user_id}, Password length: {len(password) if password else 0}")
        
        # Create server connection
        logging.debug(f"[AD] Connecting to {app.config['LDAP_HOST']}:{app.config['LDAP_PORT']}")
        server = Server(app.config['LDAP_HOST'], port=app.config['LDAP_PORT'], use_ssl=app.config['LDAP_USE_SSL'], get_info=ALL)
        
        # Use UPN format (same as login) with SIMPLE authentication
        user_bind = f"{user_id}@{app.config['LDAP_DOMAIN']}.COM"
        logging.debug(f"[AD] Using UPN bind: {user_bind}")
        
        conn = Connection(
            server,
            user=user_bind,
            password=password,
            authentication=SIMPLE
        )

        if not conn.bind():
            logging.error(f"[AD] Failed to bind to AD server: {conn.result}")
            return []

        logging.debug("[AD] Successfully bound to Active Directory")

        # Modify the search filter to include the search query if provided
        search_filter = '(objectClass=user)'
        if search_query:
            search_filter = f'(&(objectClass=user)(|(sAMAccountName=*{search_query}*)(displayName=*{search_query}*)))'

        logging.debug(f"[AD] Using search base: {app.config['LDAP_BASE_DN']}")
        logging.debug(f"[AD] Using search filter: {search_filter}")

        conn.search(app.config['LDAP_BASE_DN'], search_filter, attributes=['sAMAccountName', 'displayName', 'department', 'mail'])
        
        logging.debug(f"[AD] Search returned {len(conn.entries)} entries")
        
        for entry in conn.entries:
            username = str(entry.sAMAccountName) if 'sAMAccountName' in entry else ''
            display_name = str(entry.displayName) if 'displayName' in entry else username
            department = str(entry.department) if 'department' in entry else 'No Department'
            email = str(entry.mail) if 'mail' in entry else ''
            
            logging.debug(f"[AD] Processing entry - username: {username}, display_name: {display_name}")
            
            if username and username.lower() != user_id.lower():  # Exclude current user (case-insensitive)
                users.append({
                    'username': username,
                    'display_name': display_name,
                    'department': department,
                    'email': email
                })

        conn.unbind()
        logging.debug(f"[AD] Fetched {len(users)} users from AD")
    except Exception as e:
        logging.error(f"[AD] Exception occurred: {str(e)}", exc_info=True)
    
    logging.debug(f"[AD] Returning {len(users)} users")
    return users


@app.route('/chat')
@login_required
def chat():
    if 'user_id' not in session:
        logging.error("User ID not found in session. Redirecting to login.")
        return redirect(url_for('login'))

    logging.debug(f"Loading chat page for user: {session['user_id']}")
    ad_users = get_ad_users()
    return render_template('chat.html', users=ad_users)

@app.route('/api/chat/users/search', methods=['GET'])
def api_chat_users_search():
    """API endpoint to search for users in Active Directory using UPN format"""
    print("[API] /api/chat/users/search called")
    
    # Check if user is logged in
    if 'user_id' not in session:
        print("[API] User not logged in")
        return jsonify([]), 401
    
    search_term = request.args.get('search', '').strip()
    print(f"[API] Search term: '{search_term}'")
    
    try:
        # Get stored password from session
        password = session.get('password')
        username = session.get('user_id')
        
        print(f"[API] Session username: {username}")
        print(f"[API] Session password exists: {bool(password)}")
        
        if not password or not username:
            print("[API] Missing password or username in session")
            return jsonify([]), 401
        
        # Connect to AD using UPN format (same as login) with SIMPLE authentication
        print(f"[API] Connecting to AD: {app.config['LDAP_HOST']}:{app.config['LDAP_PORT']}")
        server = Server(app.config['LDAP_HOST'], port=app.config['LDAP_PORT'], use_ssl=app.config['LDAP_USE_SSL'], get_info=ALL)
        
        # Use UPN format like login does
        user_bind = f"{username}@{app.config['LDAP_DOMAIN']}.COM"
        print(f"[API] Using UPN bind: {user_bind}")
        
        conn = Connection(server, user=user_bind, password=password, authentication=SIMPLE)
        
        if not conn.bind():
            print(f"[API] Failed to bind with UPN: {user_bind}")
            print(f"[API] Bind result: {conn.result}")
            return jsonify([]), 401
        
        print("[API] Successfully bound to AD")
        
        # Search filter - search by givenName, sn, or sAMAccountName
        if search_term:
            search_filter = f'(&(objectClass=user)(|(givenName=*{search_term}*)(sn=*{search_term}*)(sAMAccountName=*{search_term}*)))'
        else:
            search_filter = '(objectClass=user)'
        
        print(f"[API] Search filter: {search_filter}")
        
        conn.search(
            search_base=app.config['LDAP_BASE_DN'],
            search_filter=search_filter,
            attributes=['sAMAccountName', 'givenName', 'sn', 'displayName', 'department', 'mail']
        )
        
        print(f"[API] Search returned {len(conn.entries)} entries")
        
        users = []
        for entry in conn.entries:
            try:
                username_val = str(entry.sAMAccountName).strip() if entry.sAMAccountName else 'Unknown'
                # Build display name from givenName (first name) and sn (last name)
                given_name = str(entry.givenName).strip() if entry.givenName else ''
                surname = str(entry.sn).strip() if entry.sn else ''
                if given_name and surname:
                    display_name_val = f"{given_name} {surname}"
                elif given_name:
                    display_name_val = given_name
                elif surname:
                    display_name_val = surname
                else:
                    display_name_val = str(entry.displayName).strip() if entry.displayName else username_val
                
                # Exclude current user (case-insensitive comparison)
                if username_val.lower() != username.lower():
                    users.append({
                        'username': username_val,
                        'display_name': display_name_val,
                        'department': str(entry.department).strip() if entry.department else 'No Department',
                        'email': str(entry.mail).strip() if entry.mail else ''
                    })
                    print(f"[API] Added user: {username_val} - {display_name_val}")
            except Exception as e:
                print(f"[API] Error processing entry: {str(e)}")
        
        conn.unbind()
        print(f"[API] Returning {len(users)} users")
        return jsonify(users), 200
        
    except Exception as e:
        print(f"[API] EXCEPTION: {type(e).__name__}: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify([]), 500

@app.route('/api/chat/messages/<receiver_username>', methods=['GET'])
@login_required
def get_chat_messages(receiver_username):
    """Fetch chat message history between current user and receiver"""
    sender = session.get('user_id')
    print(f"[MSG] Loading message history for {sender} <-> {receiver_username}")
    
    try:
        conn = sqlite3.connect(APP_DB)
        c = conn.cursor()
        
        # Get all messages between these two users (in both directions)
        c.execute("""
            SELECT sender, receiver, message, timestamp 
            FROM messages 
            WHERE (sender = ? AND receiver = ?) OR (sender = ? AND receiver = ?)
            ORDER BY timestamp ASC
        """, (sender, receiver_username, receiver_username, sender))
        
        messages = []
        for row in c.fetchall():
            messages.append({
                'sender': row[0],
                'receiver': row[1],
                'message': row[2],
                'timestamp': row[3]
            })
        
        conn.close()
        print(f"[MSG] Loaded {len(messages)} messages for {sender} with {receiver_username}")
        return jsonify(messages), 200
        
    except Exception as e:
        logging.error(f"[MSG] Error loading messages: {str(e)}")
        return jsonify([]), 500

@app.route('/api/chat/unread', methods=['GET'])
@login_required
def get_unread_messages():
    """Get all unread messages for the current user"""
    receiver = session.get('user_id')
    print(f"[UNREAD] Fetching unread messages for {receiver}")
    
    try:
        conn = sqlite3.connect(APP_DB)
        c = conn.cursor()
        
        # Get all unread messages for this user
        c.execute("""
            SELECT id, sender, receiver, message, timestamp 
            FROM messages 
            WHERE receiver = ? AND read = 0
            ORDER BY timestamp ASC
        """, (receiver,))
        
        unread_messages = []
        for row in c.fetchall():
            unread_messages.append({
                'id': row[0],
                'sender': row[1],
                'receiver': row[2],
                'message': row[3],
                'timestamp': row[4]
            })
        
        conn.close()
        print(f"[UNREAD] Found {len(unread_messages)} unread messages for {receiver}")
        return jsonify(unread_messages), 200
        
    except Exception as e:
        logging.error(f"[UNREAD] Error loading unread messages: {str(e)}")
        return jsonify([]), 500

@app.route('/api/tickets/pending-count', methods=['GET'])
@login_required
def get_pending_tickets_count():
    """Get count of open/in-progress tickets assigned to user or their department"""
    user_id = session.get('user_id', '')
    user_dept = session.get('department', '')
    
    try:
        conn = get_db()
        c = conn.cursor()
        
        # Count tickets that are:
        # 1. Assigned directly to the user, OR
        # 2. Assigned to user's department and not yet assigned to anyone
        # AND status is not closed
        c.execute("""
            SELECT COUNT(*) FROM tickets 
            WHERE status != 'closed' 
            AND (
                (assigned_to_username IS NOT NULL AND LOWER(assigned_to_username) = LOWER(?))
                OR 
                (assigned_department = ? AND (assigned_to_username IS NULL OR assigned_to_username = ''))
            )
        """, (user_id, user_dept))
        
        count = c.fetchone()[0]
        conn.close()
        
        return jsonify({'count': count}), 200
        
    except Exception as e:
        logging.error(f"[TICKETS] Error getting pending count: {str(e)}")
        return jsonify({'count': 0}), 500

@app.route('/api/chat/mark-read/<int:message_id>', methods=['POST'])
@login_required
def mark_message_read(message_id):
    """Mark a message as read"""
    receiver = session.get('user_id')
    
    try:
        conn = sqlite3.connect(APP_DB)
        c = conn.cursor()
        
        # Mark message as read (only if receiver matches)
        c.execute(
            "UPDATE messages SET read = 1 WHERE id = ? AND receiver = ?",
            (message_id, receiver)
        )
        conn.commit()
        conn.close()
        
        print(f"[UNREAD] Marked message {message_id} as read for {receiver}")
        return jsonify({'status': 'success'}), 200
        
    except Exception as e:
        logging.error(f"[UNREAD] Error marking message as read: {str(e)}")
        return jsonify({'status': 'error'}), 500

@app.route('/api/chat/mark-all-read/<sender_username>', methods=['POST'])
@login_required
def mark_all_read_from_sender(sender_username):
    """Mark all messages from a sender as read"""
    receiver = session.get('user_id')
    
    try:
        conn = sqlite3.connect(APP_DB)
        c = conn.cursor()
        
        # Mark all messages from this sender as read
        c.execute(
            "UPDATE messages SET read = 1 WHERE sender = ? AND receiver = ? AND read = 0",
            (sender_username, receiver)
        )
        count = c.rowcount
        conn.commit()
        conn.close()
        
        print(f"[UNREAD] Marked {count} messages from {sender_username} as read for {receiver}")
        return jsonify({'status': 'success', 'count': count}), 200
        
    except Exception as e:
        logging.error(f"[UNREAD] Error marking messages as read: {str(e)}")
        return jsonify({'status': 'error'}), 500

@app.route('/api/debug/session')
@login_required
def debug_session():
    """Debug endpoint to check session data"""
    return jsonify({
        'user_id': session.get('user_id'),
        'username': session.get('username'),
        'role': session.get('role'),
        'password_stored': 'password' in session,
        'password_length': len(session.get('password', ''))
    })

@app.route('/api/debug/ad-connection')
@login_required
def debug_ad_connection():
    """Debug endpoint to test AD connection using UPN format"""
    result = {
        'status': 'unknown',
        'message': '',
        'config': {
            'host': app.config['LDAP_HOST'],
            'port': app.config['LDAP_PORT'],
            'base_dn': app.config['LDAP_BASE_DN'],
            'domain': app.config['LDAP_DOMAIN']
        },
        'session': {
            'user_id': session.get('user_id'),
            'password_stored': 'password' in session
        }
    }
    
    try:
        if 'user_id' not in session or 'password' not in session:
            result['status'] = 'error'
            result['message'] = 'User ID or password not in session'
            return jsonify(result), 400
        
        user_id = session.get('user_id')
        password = session.get('password')
        
        logging.debug(f"[DEBUG] Attempting AD connection with user: {user_id}")
        
        server = Server(app.config['LDAP_HOST'], port=app.config['LDAP_PORT'], use_ssl=app.config['LDAP_USE_SSL'], get_info=ALL)
        # Use UPN format (same as login)
        user_bind = f"{user_id}@{app.config['LDAP_DOMAIN']}.COM"
        
        conn = Connection(
            server,
            user=user_bind,
            password=password,
            authentication=SIMPLE
        )

        if not conn.bind():
            result['status'] = 'error'
            result['message'] = f'Failed to bind: {conn.result}'
            return jsonify(result), 400
        
        # Try to search for users
        conn.search(app.config['LDAP_BASE_DN'], '(objectClass=user)', attributes=['sAMAccountName'], size_limit=10)
        
        user_count = len(conn.entries)
        conn.unbind()
        
        result['status'] = 'success'
        result['message'] = f'Connected successfully. Found {user_count} users'
        result['user_count'] = user_count
        
    except Exception as e:
        result['status'] = 'error'
        result['message'] = str(e)
        logging.error(f"[DEBUG] AD connection error: {str(e)}", exc_info=True)
    
    return jsonify(result)


@socketio.on('connect')
def handle_connect():
    username = session.get('user_id')
    if username:
        connected_users[username] = request.sid
        logging.debug(f"User connected: {username}, SID: {request.sid}")

@socketio.on('disconnect')
def handle_disconnect():
    username = session.get('user_id')
    if username in connected_users:
        del connected_users[username]
        logging.debug(f"User disconnected: {username}")

# Send message using domain usernames for routing
@socketio.on('send_message')
def handle_send_message(data):
    sender = session.get('user_id')  # Always use the logged-in user as sender
    receiver = data.get('receiver', '').strip()
    message = data.get('message', '').strip()
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    logging.debug(f"[CHAT] Message received: sender={sender}, receiver={receiver}")

    # Validate inputs
    if not receiver or not message:
        logging.warning(f"[CHAT] Invalid message data: receiver={receiver}, message length={len(message)}")
        emit('error', {'message': 'Invalid message data'}, room=request.sid)
        return

    if sender == receiver:
        logging.warning(f"[CHAT] User {sender} attempted to send message to themselves")
        emit('error', {'message': 'Cannot send message to yourself'}, room=request.sid)
        return

    # Verify receiver exists in AD using case-insensitive comparison
    try:
        # Quick AD check: try to get users and verify receiver exists
        ad_users = get_ad_users()
        receiver_exists = any(user['username'].lower() == receiver.lower() for user in ad_users)
        
        if not receiver_exists:
            logging.warning(f"[CHAT] Receiver {receiver} not found in AD")
            emit('error', {'message': f'User {receiver} not found'}, room=request.sid)
            return
        
        # Normalize receiver to match AD format (get the actual username case from AD)
        actual_receiver = next(user['username'] for user in ad_users if user['username'].lower() == receiver.lower())
        receiver = actual_receiver
        
        logging.debug(f"[CHAT] Receiver normalized to: {receiver}")
    except Exception as e:
        logging.error(f"[CHAT] Error validating receiver: {str(e)}")
        emit('error', {'message': 'Error validating user'}, room=request.sid)
        return

    # Store message in database
    try:
        conn = sqlite3.connect(APP_DB)
        c = conn.cursor()
        c.execute(
            "INSERT INTO messages (sender, receiver, message, timestamp) VALUES (?, ?, ?, ?)",
            (sender, receiver, message, timestamp)
        )
        conn.commit()
        conn.close()
        logging.debug(f"[CHAT] Message stored in DB: {sender} -> {receiver}")
    except Exception as e:
        logging.error(f"[CHAT] Error storing message: {str(e)}")
        emit('error', {'message': 'Error saving message'}, room=request.sid)
        return

    # Send the message to the recipient if they are connected
    message_data = {
        'sender': sender,
        'receiver': receiver,
        'message': message,
        'timestamp': timestamp
    }
    
    # Notify recipient if they're connected (real-time delivery)
    if receiver in connected_users:
        logging.debug(f"[CHAT] Sending real-time message to connected user: {receiver}")
        emit('receive_message', message_data, room=connected_users[receiver])
    else:
        logging.debug(f"[CHAT] Receiver {receiver} is offline. Message stored in DB for later retrieval.")

    # Send confirmation to sender
    emit('message_sent', {'status': 'success', 'message': 'Message sent'}, room=request.sid)
    emit('receive_message', message_data, room=request.sid)


# =================== CHATBOT API ====================
@app.route('/api/chatbot/message', methods=['POST'])
@login_required
def chatbot_message():
    """Process a message and return chatbot response"""
    try:
        data = request.get_json()
        user_message = data.get('message', '').strip()
        
        if not user_message:
            return jsonify({'error': 'Message is required'}), 400
        
        chatbot = get_chatbot()
        response = chatbot.get_response(user_message)
        
        return jsonify({
            'response': response,
            'timestamp': datetime.now().strftime('%H:%M')
        }), 200
        
    except Exception as e:
        logging.error(f"[CHATBOT] Error processing message: {str(e)}")
        return jsonify({'error': 'Failed to process message'}), 500

@app.route('/api/chatbot/quick-actions', methods=['GET'])
@login_required
def chatbot_quick_actions():
    """Get quick action suggestions for chatbot"""
    try:
        chatbot = get_chatbot()
        actions = chatbot.get_quick_actions()
        return jsonify({'actions': actions}), 200
    except Exception as e:
        logging.error(f"[CHATBOT] Error getting quick actions: {str(e)}")
        return jsonify({'actions': []}), 500

@app.route('/chatbot')
@login_required
def chatbot_page():
    """Render the chatbot page"""
    return render_template('chatbot.html',
                           username=session.get('username'),
                           role=session.get('role'),
                           features_list=session.get('features', '').split(','))

# =================== RUN ====================
if __name__ == '__main__':
    host = os.getenv('FLASK_HOST', '0.0.0.0')
    port = int(os.getenv('FLASK_PORT', 5000))
    debug = os.getenv('FLASK_DEBUG', 'True').lower() == 'true'
    socketio.run(app, host=host, port=port, debug=debug)
